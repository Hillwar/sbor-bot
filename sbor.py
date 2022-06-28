from openpyxl import Workbook, load_workbook
from people import Person, PersonRole, Service, Squad
from sbor_data_parser import get_sbor, save_sbor
from tools import Tools


class Sbor:
    def __init__(self, excel_path):
        self.__excel_path = excel_path
        self.load()

    def get_people_count(self):
        return len(self.__people)

    def get_squads_count(self):
        return len(self.__squads)

    def get_services_count(self):
        return len(self.__services)

    def get_commanders_count(self):
        return len(self.__commanders)

    def get_roles_count(self):
        return len(self.__roles)

    def get_person(self, id):
        return self.__people[id - 1]

    def get_squad(self, id):
        return self.__squads[id - 1]

    def get_service(self, id):
        return self.__services[id - 1]

    def get_role(self, id):
        return self.__roles[id - 1]

    def get_squad_info_people(self, squad):
        vozhatiy = self.get_person(squad.vozhatiy_id)
        komsorg = self.get_person(squad.komsorg_id)
        commander = self.get_person(self.get_commander_by_squad(squad.id).commander_id)
        return vozhatiy, komsorg, commander

    def get_squad_info_full(self, squad):
        vozhatiy, komsorg, commander = self.get_squad_info_people(squad)
        separator_middle = '\n'
        separator_end = '\n\n'
        return '*Отряд \'' + squad.name + '\'*' + separator_end + \
               '*Вожатый*' + separator_middle + self.get_person_info_one_line(vozhatiy,
                                                                              Person.Info.Full) + separator_end + \
               '*Комсорг*' + separator_middle + self.get_person_info_one_line(komsorg,
                                                                              Person.Info.Full) + separator_end + \
               '*ДКО*' + separator_middle + self.get_person_info_one_line(commander, Person.Info.Full)

    def get_squad_info_compact(self, squad):
        vozhatiy, komsorg, commander = self.get_squad_info_people(squad)
        separator_middle = ' - '
        separator_end = '\n'
        return '*Отряд \'' + squad.name + '\'*' + separator_end + \
               'Вожатый' + separator_middle + self.get_person_info_one_line(vozhatiy,
                                                                            Person.Info.Compact) + separator_end + \
               'Комсорг' + separator_middle + self.get_person_info_one_line(komsorg,
                                                                            Person.Info.Compact) + separator_end + \
               'ДКО' + separator_middle + self.get_person_info_one_line(commander, Person.Info.Compact)

    def get_squad_info_with_people(self, squad):
        people = self.get_people(Person.Filter.squad_id(squad.id), Person.Sort.surname)
        return self.get_squad_info_full(squad) + '\n\n' + self.get_people_info_grouped_by_roles(people)

    def get_person_info_one_line(self, person, info=Person.Info.Compact, index=None, name_first=False):
        id = None
        squad_id = None
        phone_number = None
        if info == Person.Info.Full:
            phone_number = person.get_phone_number()
        elif info == Person.Info.Debug:
            index = None
            id = person.id
            squad_id = person.squad_id

        person_info = person.get_full_name(name_first)
        if phone_number:
            person_info += ' ' + phone_number
        if squad_id:
            person_info = Tools.get_index(squad_id, self.get_squads_count(), '`(o', ')` ') + person_info
        if id:
            person_info = Tools.get_index(id, self.get_people_count(), '`<i', '>` ') + person_info
        if index:
            person_info = Tools.get_index(index, self.get_people_count(), '`[', ']` ') + person_info
        return person_info

    def get_person_info(self, person, info=Person.Info.Compact, name_first=False):
        id = None
        squad_id = None
        phone_number = person.get_phone_number()

        if info == Person.Info.Full:
            squad_id = person.squad_id
        elif info == Person.Info.Debug:
            id = person.id
            squad_id = person.squad_id

        person_info = '*' + person.get_full_name(name_first) + '*\n'
        if id:
            person_info += 'ID: `{}`\n'.format(id)
        if squad_id:
            person_info += 'Отряд: \'{}\'\n'.format(self.get_squad(squad_id).name)
        if phone_number:
            person_info += 'Телефон: {}\n'.format(phone_number)
        return person_info

    def get_service_info(self, service):
        if not service.supervisor_id:
            return '*' + service.name + '*\n' + 'Пока нет ответственного'

        supervisor = self.get_person(service.supervisor_id)
        return '*' + service.name + '*\n' + self.get_person_info_one_line(supervisor, Person.Info.Full)

    def get_commander_info(self, commander):
        squad_id = commander.commander_squad_id
        is_dks = squad_id == None
        nickname = '*Дежурный Командир Сбора*' if is_dks else 'ДКО _\'' + self.get_squad(squad_id).name + '\'_'
        info = Person.Info.Compact if is_dks else Person.Info.Full
        commander = self.get_person(commander.commander_id)

        commander_info = nickname + '\n'
        commander_info += self.get_person_info_one_line(commander, info)
        if is_dks:
            commander_info += ' ' + self.__info.get_dks_number()

        return commander_info

    def get_commander_by_squad(self, squad_id):
        squad_commanders = list(filter(lambda commander: commander.commander_squad_id == squad_id, self.__commanders))
        return squad_commanders[0]

    def get_people(self, people_filter=None, people_sort=None):
        result = list(self.__people)
        if people_filter:
            result = list(filter(people_filter, result))
        if people_sort:
            result.sort(key=people_sort)
        return result

    def get_people_grouped_by_roles(self, people):
        people_grouped = dict()
        for role_index in range(1, self.get_roles_count() + 1):
            role = self.get_role(role_index)
            people_grouped[role.id] = Person.filter(people, Person.Filter.role_id(role.id))

        return people_grouped

    def get_people_info_grouped_by_roles(self, people, info = Person.Info.Compact):
        people_info = ''
        people_grouped = self.get_people_grouped_by_roles(people)
        for role_index in range(1, self.get_roles_count() + 1):
            role = self.get_role(role_index)
            people_info += '\n\n' if people_info else ''
            people_info += '*{}*\n'.format(role.plural)
            people_info += self.get_people_info(people_grouped[role.id], info)

        return people_info

    def get_squads_info(self):
        squad_info = ''
        for squad in self.__squads:
            squad_info += '\n\n' if squad_info else ''
            squad_info += self.get_squad_info_compact(squad)
        return squad_info

    def get_supervisors_info(self):
        supervisors_info = ''
        for supervisor in self.__supervisors:
            supervisors_info += '\n\n' if supervisors_info else ''
            supervisors_info += self.get_service_info(supervisor)
        return supervisors_info

    def get_sbor_info(self):
        sbor_info = '*СБОР {}*\n\n'.format(self.__info.number)
        sbor_info += self.get_supervisors_info() + '\n\n'
        sbor_info += '*Адрес*\n[{}]({})\n\n'.format(self.__info.adress, self.__info.location_link)
        sbor_info += '*Группа Вконтакте*\n[Ссылка]({})'.format(self.__info.vk_link)
        return sbor_info

    def get_services_info(self):
        service_info = ''
        for service in self.__services:
            service_info += '\n\n' if service_info else ''
            service_info += self.get_service_info(service)
        return service_info

    def get_commanders_info(self):
        commanders_info = ''
        for commander in self.__commanders:
            commanders_info += '\n\n' if commanders_info else ''
            commanders_info += self.get_commander_info(commander)
        return commanders_info

    def get_people_info_by_ids(self, people_ids, info=Person.Info.Compact):
        people = []
        for id in people_ids:
            people.append(self.get_person(id))

        return self.get_people_info(people, info)

    def get_people_info(self, people, info=Person.Info.Compact, name_first=False):
        people_info = ''
        index = 1
        for person in people:
            people_info += '\n' if people_info else ''
            people_info += self.get_person_info_one_line(person, index=index, info=info, name_first=name_first)
            index += 1
        return people_info

    def get_people_grouped_by(self, groupby, range, sort, namegetter, info=Person.Info.Compact):
        grouped_people = {}
        for index in range:
            grouped_people[index] = self.get_people(groupby(index), sort)
        return grouped_people

    def get_people_list_info(self, sort, people, info=Person.Info.Compact, name_first=False):
        people.sort(key=sort)
        return '*Список участников*\n' + self.get_people_info(people, info, name_first)

    def get_all_people_info(self, sort, info=Person.Info.Compact, name_first=False):
        people = list(self.__people)
        return self.get_people_list_info(sort, people, info, name_first)

    def get_squad_people_info(self, sort, info=Person.Info.Compact, name_first=False):
        people = self.get_people(lambda person: person.squad_id)
        return self.get_people_list_info(sort, people, info, name_first)

    def __find_people_by_key(self, key):
        if key.isdigit():
            id = int(key)
            return self.get_people(Person.Filter.id(id))

        key = key.lower()
        people = self.get_people(Person.Filter.name(key))
        people.extend(self.get_people(Person.Filter.surname(key)))
        return people

    def find_people(self, keys):
        people = None
        for key in keys:
            if people:
                people = people & set(self.__find_people_by_key(key))
            else:
                people = set(self.__find_people_by_key(key))
        return people

    def edit_commanders(self, new_main_commander_id, new_commanders_ids):
        if not new_main_commander_id or len(new_commanders_ids) != self.get_squads_count():
            return False, "Нужно передать 1 ДКС и {} ДКО. Вы передали {} ДКС и {} ДКО.".format(self.get_squads_count(),
                                                                                               1 if new_main_commander_id else 0,
                                                                                               len(new_commanders_ids))

        def in_range(id):
            return id > 0 and id <= self.get_people_count()

        if (not in_range(new_main_commander_id)):
            return False, "ID должен быть числом больше 0 и меньше {}.".format(self.get_people_count() + 1)

        squads_with_new_commanders = {}
        unique_commanders = set()
        unique_commanders.add(new_main_commander_id)
        for new_commander_id in new_commanders_ids:
            if (not in_range(new_commander_id)):
                return False, "ID должен быть числом больше 0 и меньше {}.".format(self.get_people_count() + 1)

            new_commander_squad = self.get_person(new_commander_id).squad_id
            if new_commander_id in unique_commanders:
                return False, "ID не должны повторяться."
            if new_commander_squad in squads_with_new_commanders:
                return False, "{}\n\nДКО должны быть членами разных отрядов!".format(
                    self.get_people_info_by_ids(new_commanders_ids, Person.Info.Debug))

            squads_with_new_commanders[new_commander_squad] = new_commander_id
            unique_commanders.add(new_commander_id)

        for i in range(self.get_commanders_count()):
            if self.__commanders[i].commander_squad_id:
                self.__commanders[i].commander_id = squads_with_new_commanders[self.__commanders[i].commander_squad_id]
            else:
                self.__commanders[i].commander_id = new_main_commander_id

        return True, ""

    def save(self):
        workbook = load_workbook(self.__excel_path)
        save_sbor(workbook, self.__excel_path, self.__commanders)

    def load(self):
        workbook = load_workbook(self.__excel_path, data_only=True)
        self.__people, self.__squads, self.__commanders, self.__services, self.__roles, self.__supervisors, self.__info = get_sbor(
            workbook)
