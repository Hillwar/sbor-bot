from openpyxl import Workbook, load_workbook
from parse.solovki_parser import save_solovki
from people import Person, PersonRole, Service, Squad
from parse.solovki_parser import get_solovki, save_solovki
from tools import Tools

class Solovki:
    def __init__(self, excel_path):
        self.__excel_path = excel_path
        self.load()

    def get_people_count(self):
        return len(self.__people)

    def get_squads_count(self):
        return len(self.__squads)

    def get_person(self, id):
        return self.__people[int(id) - 1]

    def get_squad(self, id):
        return self.__squads[id - 1]

    def get_squad_info_people(self, squad):
        supervisor = self.get_person(squad.supervisor_id)
        return supervisor

    def get_squad_info_full(self, squad):
        supervisor = self.get_squad_info_people(squad)
        separator_middle = '\n'
        separator_end = '\n\n'
        return '*Тройка \'' + squad.name + '\'*' + separator_end + \
               '*Ответсвенный*' + separator_middle + self.get_person_info_one_line(supervisor, Person.Info.Full)

    def get_squad_info_with_people(self, squad):
        people = self.get_people(Person.Filter.squad_id(squad.id), Person.Sort.surname)
        return self.get_squad_info_full(squad) + '\n\n*Состав*\n' + self.get_people_info(people)

    def get_person_info_one_line(self, person, info=Person.Info.Compact, index=None, name_first=False):
        id = None
        squad_id = None
        phone_number = None
        if info == Person.Info.Full:
            phone_number = person.get_phone_number()
        elif info == Person.Info.Debug:
            index = None
            id = person.id
            squad_id = person.squad_id

        person_info = person.get_full_name(name_first)
        if phone_number:
            person_info += ' ' + phone_number
        if squad_id:
            person_info = Tools.get_index(squad_id, self.get_squads_count(), '`(s', ')` ') + person_info
        if id:
            person_info = Tools.get_index(id, self.get_people_count(), '`<i', '>` ') + person_info
        if index:
            person_info = Tools.get_index(index, self.get_people_count(), '`[', ']` ') + person_info
        return person_info

    def get_person_info(self, person, info=Person.Info.Compact, name_first=False):
        id = None
        squad_id = None
        phone_number = person.get_phone_number()

        if info == Person.Info.Full:
            squad_id = person.squad_id
        elif info == Person.Info.Debug:
            id = person.id
            squad_id = person.squad_id

        person_info = '*' + person.get_full_name(name_first) + '*\n'
        if id:
            person_info += 'ID: `{}`\n'.format(id)
        if squad_id:
            person_info += 'Тройка: \'{}\'\n'.format(self.get_squad(squad_id).name)
        if phone_number:
            person_info += 'Телефон: {}\n'.format(phone_number)
        return person_info

    def get_service_info(self, service):
        if not service.supervisor_id:
            return '*' + service.name + '*\n' + 'Пока нет ответственного'

        supervisor = self.get_person(service.supervisor_id)
        return '*' + service.name + '*\n' + self.get_person_info_one_line(supervisor, Person.Info.Full)

    def get_commander_info(self, commander):
        nickname = '*Дежурный Командир Лагеря*'
        info = Person.Info.Full
        commander = self.get_person(commander.commander_id)
        commander_info = nickname + '\n'
        commander_info += self.get_person_info_one_line(commander, info)
        return commander_info

    def get_people(self, people_filter=None, people_sort=None):
        result = list(self.__people)
        if people_filter:
            result = list(filter(people_filter, result))
        if people_sort:
            result.sort(key=people_sort)
        return result

    def get_supervisors_info(self):
        supervisors_info = ''
        for supervisor in self.__supervisors:
            supervisors_info += '\n\n' if supervisors_info else ''
            supervisors_info += self.get_service_info(supervisor)
        return supervisors_info

    def get_solovki_info(self):
        sbor_info = '*СОЛОВКИ {}*\n\n'.format(self.__info.number)
        sbor_info += self.get_supervisors_info() + '\n\n'

        if self.__info.adress and self.__info.location_link:
            sbor_info += '*Адрес*\n[{}]({})\n\n'.format(self.__info.adress, self.__info.location_link)
        if self.__info.vk_link:
            sbor_info += '*Группа Вконтакте*\n[Ссылка]({})'.format(self.__info.vk_link)
        if self.__info.tg_chat_link:
            sbor_info += '*Чат Telegram*\n[Ссылка]({})'.format(self.__info.tg_chat_link)
        return sbor_info

    def get_services_info(self):
        service_info = ''
        for service in self.__services:
            service_info += '\n\n' if service_info else ''
            service_info += self.get_service_info(service)
        return service_info

    def get_commanders_info(self):
        commanders_info = ''
        for commander in self.__commanders:
            commanders_info += '\n\n' if commanders_info else ''
            commanders_info += self.get_commander_info(commander)
        return commanders_info

    def get_people_info_by_ids(self, people_ids, info=Person.Info.Compact):
        people = []
        for id in people_ids:
            people.append(self.get_person(id))

        return self.get_people_info(people, info)

    def get_people_info(self, people, info=Person.Info.Compact, name_first=False):
        people_info = ''
        index = 1
        for person in people:
            people_info += '\n' if people_info else ''
            people_info += self.get_person_info_one_line(person, index=index, info=info, name_first=name_first)
            index += 1
        return people_info

    def get_people_grouped_by(self, groupby, range, sort, namegetter, info=Person.Info.Compact):
        grouped_people = {}
        for index in range:
            grouped_people[index] = self.get_people(groupby(index), sort)
        return grouped_people

    def get_people_list_info(self, sort, people, info=Person.Info.Compact, name_first=False):
        people.sort(key=sort)
        return '*Список участников*\n' + self.get_people_info(people, info, name_first)

    def get_all_people_info(self, sort, info=Person.Info.Compact, name_first=False):
        people = list(self.__people)
        return self.get_people_list_info(sort, people, info, name_first)

    def get_squad_people_info(self, sort, info=Person.Info.Compact, name_first=False):
        people = self.get_people(lambda person: person.squad_id)
        return self.get_people_list_info(sort, people, info, name_first)

    def __find_people_by_key(self, key):
        if key.isdigit():
            id = int(key)
            return self.get_people(Person.Filter.id(id))

        key = key.lower()
        people = self.get_people(Person.Filter.name(key))
        people.extend(self.get_people(Person.Filter.surname(key)))
        return people

    def find_people(self, keys):
        people = None
        for key in keys:
            if people:
                people = people & set(self.__find_people_by_key(key))
            else:
                people = set(self.__find_people_by_key(key))
        return people

    def edit_commanders(self, new_main_commander_id, new_commanders_ids):
        if not new_main_commander_id:
            return False, "Нужно передать 1 ДКЛ. Вы не передали никого"

        def in_range(id):
            return id > 0 and id <= self.get_people_count()

        if (not in_range(new_main_commander_id)):
            return False, "ID должен быть числом больше 0 и меньше {}.".format(self.get_people_count() + 1)

        self.__commanders[0].commander_id = new_main_commander_id
        return True, ""

    def save(self):
        workbook = load_workbook(self.__excel_path)
        save_solovki(workbook, self.__excel_path, self.__commanders)

    def load(self):
        workbook = load_workbook(self.__excel_path, data_only=True)
        self.__people, self.__squads, self.__commanders, self.__services, self.__supervisors, self.__info = get_solovki(
            workbook)
