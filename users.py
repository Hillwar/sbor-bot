from openpyxl import Workbook, load_workbook
from people import Admin, AdminRole
from parser import get_users, save_users, save_admins, get_admins
from tools import Tools

class Users:
    def __init__(self, document_path):
        self.__document_path = document_path
        self.load()

    def add_user(self, user):
        if user in self.__users:
            return False

        self.__users.add(user)
        return True

    def get_users(self):
        return list(self.__users)

    def save(self):
        # Возникнут проблемы с много-поточкой. Нужно написать защиту
        with open(self.__document_path, 'w') as file:
            save_users(file, self.__users)

    def load(self):
        # Возникнут проблемы с много-поточкой. Нужно написать защиту
        with open(self.__document_path, 'r') as file:
            self.__users = get_users(file)


class Admins:
    def __init__(self, excel_path):
        self.__excel_path = excel_path
        self.load()

    def get_admins_count(self):
        return len(self.__admins)

    def get_roles_count(self):
        return len(self.__roles)

    def get_admin(self, id):
        return self.__admins[id - 1]

    def get_role(self, id):
        return self.__roles[id - 1]

    def can_role_of_admin(self, admin, predicate):
        role =  self.get_role(admin.role_id)
        return predicate(role)

    def get_users_who_can(self, role_predicate):
        users = []
        for admin in self.__admins:
            if self.can_role_of_admin(admin, role_predicate):
                users.append(admin.telegram)

        return users

    def get_users_in_admins_list(self):
        return self.get_users_who_can(lambda role: True)

    def get_users_who_can_public_messages(self):
        return self.get_users_who_can(lambda role: role.public_messages)

    def get_users_who_can_see_ids(self):
        return self.get_users_who_can(lambda role: role.see_ids)

    def get_users_who_can_edit_something(self):
        return self.get_users_who_can(lambda role: role.edit_commanders or role.edit_timetable or role.manage_admins)

    def get_users_who_can_edit_commanders(self):
        return self.get_users_who_can(lambda role: role.edit_commanders)

    def get_users_who_can_edit_timetable(self):
        return self.get_users_who_can(lambda role: role.edit_timetable)

    def get_users_who_can_edit_admins(self):
        return self.get_users_who_can(lambda role: role.manage_admins)

    def refresh_admins_ids(self):
        for i, admin in enumerate(self.__admins):
            admin.id = i + 1

    def add_admin(self, telegram, role_id):
        if not telegram or not isinstance(telegram, str):
            return False, 'Вы не передали Telegram админа'

        if not role_id or not isinstance(role_id, int):
            return False, 'Вы не передали ID роли'

        if not Tools.in_range(role_id, 1, self.get_roles_count()):
            return False, 'ID роли должно быть в пределах от {} до {}'.format(1, self.get_roles_count())

        if Tools.check_list_for(self.__admins, lambda admin: admin.telegram == telegram):
            return False, 'Админ с таким ником уже есть'

        admin = Admin(self.get_admins_count() + 1, telegram, role_id)
        self.__admins.append(admin)
        return True, ''

    def remove_admin(self, id):
        if not id or not isinstance(id, int):
            return False, 'Вы не передали ID админа'

        if not Tools.in_range(id, 1, self.get_admins_count()):
            return False, 'ID админа должно быть в пределах от {} до {}'.format(1, self.get_admins_count())

        del self.__admins[id - 1]
        self.refresh_admins_ids()
        return True, ''

    def edit_role_admin(self, id, new_role_id):
        if not id or not isinstance(id, int):
            return False, 'Вы не передали ID админа'

        if not new_role_id or not isinstance(new_role_id, int):
            return False, 'Вы не передали ID роли'

        if not Tools.in_range(id, 1, self.get_admins_count()):
            return False, 'ID админа должно быть в пределах от {} до {}'.format(1, self.get_admins_count())

        if not Tools.in_range(new_role_id, 1, self.get_roles_count()):
            return False, 'ID роли должно быть в пределах от {} до {}.'.format(1, self.get_roles_count())

        self.__admins[id - 1].role_id = new_role_id
        return True, ''

    def get_admin_info(self, admin):
        role = self.get_role(admin.role_id)
        return  'Telegram: `' + admin.telegram + '`\n'\
                'ID: `' + str(admin.id) + '`\n'\
                'Роль: ' + role.name

    def get_role_info(self, role):
        return  'Роль _\'' + role.name + '\'_\n'\
                'ID: `' + str(role.id) + '`\n'\
                'Может делать рассылку: ' + Tools.bool_to_russian(role.public_messages) + '\n'\
                'Может видеть ID участников: ' + Tools.bool_to_russian(role.see_ids) + '\n'\
                'Может изменять командиров: ' + Tools.bool_to_russian(role.edit_commanders) + '\n'\
                'Может изменять расписание: ' + Tools.bool_to_russian(role.edit_timetable) + '\n'\
                'Может изменять админов: ' + Tools.bool_to_russian(role.manage_admins)


    def get_roles_info(self):
        roles = list(self.__roles)
        roles_info = []

        for role in roles:
            roles_info.append(self.get_role_info(role))

        return roles_info

    def get_admins_info(self):
        admins = list(self.__admins)
        admins_info = []

        for admin in admins:
            admins_info.append(self.get_admin_info(admin))

        return admins_info

    def save(self):
        workbook = load_workbook(self.__excel_path)
        save_admins(workbook, self.__excel_path, self.__admins)

    def load(self):
        workbook = load_workbook(self.__excel_path, data_only=True)
        self.__admins, self.__roles = get_admins(workbook)